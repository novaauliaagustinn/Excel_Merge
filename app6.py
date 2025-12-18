import streamlit as st
import pandas as pd
from io import BytesIO

# ============================
# STREAMLIT SAFE CONFIG
# ============================
st.set_option('client.showErrorDetails', False)

# ============================
#  CONFIG TEMA & LAYOUT
# ============================
st.set_page_config(
    page_title="Excel Merge",
    layout="centered"
)

# ============================
# SEMBUNYIKAN PREVIEW FILE
# ============================
st.markdown("""
<style>
.uploadedFile {display: none !important;}
.stUploadedFile {display: none !important;}
</style>
""", unsafe_allow_html=True)

# ============================
# BALIK URUTAN TAMPILAN FILE
# ============================
st.markdown("""
<style>
.stFileUploader > div:nth-child(1) > div {
    display: flex;
    flex-direction: column-reverse;
}
</style>
""", unsafe_allow_html=True)

# ============================
# PERBESAR LIST FILE
# ============================
st.markdown("""
<style>
.stFileUploader > div > div {
    max-height: 450px !important;
    overflow-y: auto !important;
}
</style>
""", unsafe_allow_html=True)

# ============================
# WIDTH
# ============================
st.markdown("""
<style>
.main {
    max-width: 750px;
    margin: auto;
    padding-top: 20px;
}
</style>
""", unsafe_allow_html=True)

# ============================
# HEADER
# ============================
st.markdown("""
<div style='text-align:left; padding: 10px 0;'>
    <h1 style='margin-bottom: 0; font-size: 40px; color:#150F3D;'>
        Excel Merge
    </h1>
    <p style='font-size:16px; color:#555;'>
        This tool automatically merges multiple Excel files into a single dataset with unified formatting
    </p>
    <hr style='margin-top:15px;'>
</div>
""", unsafe_allow_html=True)

# ============================
# UPLOAD FILES
# ============================
uploaded_files = st.file_uploader(
    "Upload Excel Files",
    type=["xlsx"],
    accept_multiple_files=True
)

if uploaded_files:
    uploaded_files = list(uploaded_files)[::-1]

# ============================
# FUNGSI CARI HEADER (CEPAT)
# ============================
def find_header_row(df, max_scan=20):
    for i in range(min(len(df), max_scan)):
        if df.iloc[i].notna().sum() > 2:
            return i
    return 0

# ============================
# PROSES MERGE
# ============================
if uploaded_files:

    st.info("Uploading & Processing Files...")

    total = len(uploaded_files)
    progress = st.progress(0)
    status = st.empty()

    merged_list = []
    total_rows_all = 0

    for i, file in enumerate(uploaded_files):
        status.write(f"Processing: **{i+1}/{total}**")

        excel_data = pd.read_excel(
            file,
            sheet_name=None,
            header=None
        )

        for sheet_name, df in excel_data.items():
            header_idx = find_header_row(df)
            df.columns = df.iloc[header_idx]
            df = df.iloc[header_idx + 1:].reset_index(drop=True)

            total_rows_all += len(df)

            df["Source File"] = file.name
            df["Sheet"] = sheet_name

            merged_list.append(df)

        progress.progress((i + 1) / total)

    status.write(f"✅ Successfully processed {total} files, {total_rows_all:,} total rows")

    merged_df = pd.concat(merged_list, ignore_index=True)
    merged_df.index = merged_df.index + 1

    # ============================
    # PREVIEW (AMAN DATA BESAR)
    # ============================
    st.dataframe(merged_df.head(200), use_container_width=True)

    # ============================
    # EXPORT EXCEL (DICACHE)
    # ============================
    @st.cache_data(show_spinner=False)
    def to_excel(df):
        from openpyxl import load_workbook
        from openpyxl.styles import Font

        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Merged Data')
        writer.close()

        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active

        for cell in ws[1]:
            cell.font = Font(bold=False)

        clean_output = BytesIO()
        wb.save(clean_output)
        clean_output.seek(0)

        return clean_output.getvalue()

    # ============================
    # DOWNLOAD (NGGAK FREEZE)
    # ============================
    with st.spinner("Preparing Excel file..."):
        excel_file = to_excel(merged_df)

    st.download_button(
        label="Download Merged Excel",
        data=excel_file,
        file_name="Merged_Excel.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
